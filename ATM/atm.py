import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import ttk

CURRENT_FILE = "current_snapshot.xlsx"
MASTER_FILE = "master_atm_log.xlsx"


# ================= DATA PART =================

def read_snapshot():
    df = pd.read_excel(CURRENT_FILE)
    df['total_amount'] = df['CAS100'] + df['CAS200'] + df['CAS500']
    return df[['ATM_ID', 'total_amount']]


def initialize_master(snapshot_df):
    if not os.path.exists(MASTER_FILE):
        master_df = pd.DataFrame()
        master_df['ATM_ID'] = snapshot_df['ATM_ID']
        master_df.to_excel(MASTER_FILE, index=False)


def update_master(snapshot_df):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    master_df = pd.read_excel(MASTER_FILE)

    master_df = master_df.merge(snapshot_df, on="ATM_ID", how="left")
    master_df.rename(columns={"total_amount": timestamp}, inplace=True)

    master_df.to_excel(MASTER_FILE, index=False)
    apply_color_logic()


def apply_color_logic():
    wb = load_workbook(MASTER_FILE)
    ws = wb.active

    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")

    for row in range(2, ws.max_row + 1):
        prev = None
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)

            if prev is None:
                cell.fill = green
                prev = cell.value
            else:
                if cell.value == prev:
                    cell.fill = red
                else:
                    cell.fill = green
                    prev = cell.value

    wb.save(MASTER_FILE)


# ================= ANALYSIS PART =================

def get_down_atms():
    df = pd.read_excel(MASTER_FILE)
    result = {
        "2 hr": [],
        "4 hr": [],
        "6 hr": [],
        "8 hr": [],
        "10+ hr": []
    }

    for _, row in df.iterrows():
        values = row[1:].dropna().tolist()
        if len(values) < 2:
            continue

        streak = 0
        for i in range(len(values) - 1, 0, -1):
            if values[i] == values[i - 1]:
                streak += 1
            else:
                break

        hours = streak * 2

        if hours >= 10:
            result["10+ hr"].append(row['ATM_ID'])
        elif hours >= 8:
            result["8 hr"].append(row['ATM_ID'])
        elif hours >= 6:
            result["6 hr"].append(row['ATM_ID'])
        elif hours >= 4:
            result["4 hr"].append(row['ATM_ID'])
        elif hours >= 2:
            result["2 hr"].append(row['ATM_ID'])

    return result


# ================= GUI PART =================

def launch_gui():
    data = get_down_atms()

    root = tk.Tk()
    root.title("ATM Down-Time Monitor")
    root.geometry("400x400")

    ttk.Label(root, text="Select Down Time").pack(pady=10)

    combo = ttk.Combobox(
        root,
        values=["2 hr", "4 hr", "6 hr", "8 hr", "10+ hr"],
        state="readonly"
    )
    combo.pack()

    listbox = tk.Listbox(root, width=40, height=15)
    listbox.pack(pady=15)

    def update_list(event):
        listbox.delete(0, tk.END)
        selected = combo.get()
        for atm in data.get(selected, []):
            listbox.insert(tk.END, atm)

    combo.bind("<<ComboboxSelected>>", update_list)

    root.mainloop()


# ================= MAIN =================

def main():
    snapshot = read_snapshot()
    initialize_master(snapshot)
    update_master(snapshot)
    launch_gui()


if __name__ == "__main__":
    main()